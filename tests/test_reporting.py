from __future__ import annotations

import tempfile
from pathlib import Path
import unittest

from openpyxl import load_workbook

from lotofacil.backtest import walk_forward_backtest
from lotofacil.financial import compare_portfolios_financially
from lotofacil.portfolio import (
    compare_portfolios,
    generate_baseline_portfolio,
    generate_balanced_portfolio,
    generate_random_portfolio,
)
from lotofacil.reporting import export_analysis_xlsx, export_backtest_xlsx

from tests.helpers import valid_history, valid_history_with_prizes


class ReportingTests(unittest.TestCase):
    def test_export_analysis_xlsx_creates_expected_sheets(self) -> None:
        history = valid_history_with_prizes()
        portfolios = {
            "baseline": generate_baseline_portfolio(history, 5, seed=1, forbidden_tickets=set()),
            "aleatoria": generate_random_portfolio(5, seed=2),
            "balanceada": generate_balanced_portfolio(5, seed=3, candidate_pool=80),
        }
        evaluations = compare_portfolios(portfolios, sample_draw_count=20, seed=4)
        prize_profile, financial_evaluations = compare_portfolios_financially(
            portfolios,
            history,
            ticket_price=3.50,
            simulations=20,
            seed=4,
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "resumo.xlsx"
            export_analysis_xlsx(
                path,
                history,
                evaluations,
                portfolios,
                financial_summaries=financial_evaluations,
                prize_profile=prize_profile,
            )

            workbook = load_workbook(path, data_only=True)
            self.assertEqual(
                {
                    "analise",
                    "historico",
                    "mercado",
                    "premios",
                    "frequencias",
                    "probabilidades",
                    "estrategias",
                    "financeiro",
                    "sensibilidade",
                    "apostas_baseline",
                    "apostas_aleatoria",
                    "apostas_balanceada",
                },
                set(workbook.sheetnames),
            )
            self.assertEqual("analise", workbook.sheetnames[0])
            self.assertEqual("source_path", workbook["historico"]["A1"].value)
            self.assertEqual("strategy", workbook["financeiro"]["A1"].value)

    def test_export_backtest_xlsx_creates_expected_sheets(self) -> None:
        history = valid_history()
        report = walk_forward_backtest(history, 5, seed=7, window=2, max_test_draws=2)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "backtest.xlsx"
            export_backtest_xlsx(path, report, history)

            workbook = load_workbook(path, data_only=True)
            self.assertEqual(
                {
                    "analise",
                    "historico",
                    "mercado",
                    "premios",
                    "frequencias",
                    "probabilidades",
                    "backtest_resumo",
                    "backtest_detalhe",
                },
                set(workbook.sheetnames),
            )
            self.assertEqual("backtest_resumo", workbook.sheetnames[6])
            self.assertEqual("strategy", workbook["backtest_resumo"]["A1"].value)
